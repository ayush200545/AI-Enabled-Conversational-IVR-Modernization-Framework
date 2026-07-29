import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

def generate_agile_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agile Template"
    
    headers = ["Planned Sprint", "Actual Sprint", "US ID", "User Story Description", "MOSCOW", "Dependency", "Assignee", "Status"]
    ws.append(headers)
    
    # Add some sample data based on the project
    data = [
        ["Sprint 1", "Sprint 1", "US001", "Analyze legacy VXML system architecture", "Must Have", "None", "Intern 1", "Done"],
        ["Sprint 1", "Sprint 1", "US002", "Document integration points for ACS/BAP", "Must Have", "US001", "Intern 2", "Done"],
        ["Sprint 2", "Sprint 2", "US003", "Setup FastAPI boilerplate and ngrok", "Must Have", "None", "Intern 1", "Done"],
        ["Sprint 2", "Sprint 2", "US004", "Implement Twilio Webhook Endpoints", "Must Have", "US003", "Intern 2", "Done"],
        ["Sprint 3", "Sprint 3", "US005", "Develop AI Integration Mock Service", "Should Have", "US004", "Intern 1", "Done"],
        ["Sprint 3", "Sprint 3", "US006", "Test end-to-end voice flow", "Must Have", "US005", "Intern 2", "Done"],
        ["Sprint 4", "Sprint 4", "US007", "Complete Agile and Testing Documentation", "Must Have", "None", "Intern 1", "Done"],
        ["Sprint 4", "Sprint 4", "US008", "Prepare for final demo presentation", "Must Have", "US007", "Intern 2", "Done"]
    ]
    
    for row in data:
        ws.append(row)
        
    # Styling
    header_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    wb.save("Agile_Documentation.xlsx")


def generate_unit_test_plan():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unit Test Plan"
    
    headers = ["Test Case ID", "Module Name", "Test Description", "Expected Result", "Actual Result", "Status"]
    ws.append(headers)
    
    data = [
        ["TC001", "FastAPI App", "Verify / server status returns 200", "JSON with success message", "JSON with success message", "Pass"],
        ["TC002", "Twilio Webhook", "Verify /ivr/welcome returns valid TwiML", "Valid XML string with <Response><Gather>", "Valid XML string returned", "Pass"],
        ["TC003", "Twilio Webhook", "Verify /ivr/handle-input falls back on empty input", "TwiML asking to repeat", "TwiML asking to repeat", "Pass"],
        ["TC004", "AI Integration", "Simulate 'book flight' intent", "Response requesting destination", "Response requesting destination", "Pass"],
        ["TC005", "TwiML Utils", "Check create_twiml_say output format", "Properly formatted XML block", "Properly formatted XML block", "Pass"]
    ]
    
    for row in data:
        ws.append(row)
        
    wb.save("Unit_Test_Plan.xlsx")


def generate_defect_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defect Tracker"
    
    headers = ["Sl No", "Submitted Date", "Description", "Detected Sprint", "Assigned To", "Type Of Defect", "Action Taken", "Action Taken Date", "Status(Open/Closed)", "Remarks"]
    ws.append(headers)
    
    data = [
        [1, "2026-03-01", "Ngrok webhook URL giving 404 in Twilio Console", "Sprint 2", "Backend Team", "Configuration", "Updated ngrok URL in Twilio console voice settings", "2026-03-01", "Closed", "Verified successfully"],
        [2, "2026-03-10", "TwiML Gather timeout is too short for long queries", "Sprint 3", "Backend Team", "Functional", "Increased timeout to 5 seconds in create_twiml_gather", "2026-03-11", "Closed", "Latency improved"],
        [3, "2026-03-18", "AI layer returns 500 when SpeechResult is empty", "Sprint 3", "Integration", "Logic", "Added fallback block to handle None values", "2026-03-19", "Closed", "Handled gracefully"],
        [4, "2026-03-22", "Missing language parameter in text-to-speech causing accent issues", "Sprint 4", "UI/UX", "UI", "Added language='en-IN' to Say and Gather tags", "2026-03-22", "Closed", "Pronunciation fixed"]
    ]
    
    for row in data:
        ws.append(row)
        
    header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        
    wb.save("Defect_Tracker.xlsx")

if __name__ == "__main__":
    generate_agile_template()
    generate_unit_test_plan()
    generate_defect_tracker()
    print("Excel files generated successfully.")
